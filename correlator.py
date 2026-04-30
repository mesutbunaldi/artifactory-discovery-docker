"""
correlator.py v4
4 kaynaktan veri birleştirme + Excel rapor:
  1. Ansible roles
  2. Jenkins config.xml
  3. Jenkins console log
  4. ⭐ Servis reposu manifest dosyaları — EN GÜÇLÜ KANIT
"""
import json
import re
from pathlib import Path
from collections import Counter
from typing import Dict, List, Set
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def detect_platform(url: str) -> str:
    u = (url or "").lower()
    if "bitbucket" in u: return "Bitbucket"
    if "gitlab" in u: return "GitLab"
    if "github" in u: return "GitHub"
    if "azure" in u or "visualstudio" in u: return "Azure Repos"
    return "Unknown"


def normalize_tool_name(tool: str) -> str:
    """Aynı tool'un farklı isimlerini eşitle (Maven vs Maven Wrapper, npm vs npm/yarn)."""
    t = tool.lower().strip()
    # Maven varyantları
    if t in ("maven", "maven wrapper", "maven/gradle"):
        return "maven"
    if t in ("gradle", "gradle wrapper", "gradle (kotlin dsl)"):
        return "gradle"
    # .NET varyantları
    if t in ("dotnet cli", "msbuild", "msbuild/dotnet", "nuget", "nuget push", "nuget (legacy)"):
        return "dotnet"
    # Node varyantları
    if t in ("npm", "yarn", "pnpm", "npm/yarn"):
        return "npm-family"
    # Python varyantları
    if t in ("poetry", "poetry/pip", "pip", "pip/poetry", "setuptools", "twine", "pip (editable)"):
        return "python-build"
    # Docker
    if t in ("docker", "docker build", "docker push", "docker compose"):
        return "docker"
    # Helm
    if t.startswith("helm"):
        return "helm"
    # Go
    if t.startswith("go "):
        return "go"
    # Cargo
    if t == "cargo":
        return "cargo"
    return t


def tools_intersect(a: Set[str], b: Set[str]) -> bool:
    """İki tool seti normalize edildikten sonra kesişiyor mu?"""
    if not a or not b:
        return False
    norm_a = {normalize_tool_name(t) for t in a}
    norm_b = {normalize_tool_name(t) for t in b}
    return bool(norm_a & norm_b)


def reconcile_with_manifest(role_tools, config_tools, log_tools,
                            manifest_tools, jenkinsfile_tools):
    """
    Manifest = en güçlü kanıt.
    
    Confidence:
    - high     = manifest + en az 1 CI kaynağı uzlaşıyor
    - conflict = manifest var ama hiç CI ile uzlaşmıyor (önemli!)
    - medium   = manifest yok ama CI kaynakları uzlaşıyor
    - low      = tek kaynak
    - no-data  = hiç veri yok
    """
    ci_union = role_tools | config_tools | log_tools
    all_union = ci_union | manifest_tools | jenkinsfile_tools
    
    if not all_union:
        return {
            "primary_tool": "",
            "all_tools": "",
            "confidence": "no-data",
            "conflict_detail": "Hiçbir kaynaktan build tool tespit edilemedi",
        }
    
    # MANIFEST VAR
    if manifest_tools:
        agrees_with_ci = (tools_intersect(manifest_tools, ci_union)
                          or tools_intersect(manifest_tools, jenkinsfile_tools))
        
        if agrees_with_ci:
            # Hangi tool'larda uzlaşıyor — normalized eşleşme
            agreed = set()
            for mt in manifest_tools:
                mn = normalize_tool_name(mt)
                for ct in ci_union | jenkinsfile_tools:
                    if normalize_tool_name(ct) == mn:
                        agreed.add(mt)
                        break
            return {
                "primary_tool": ", ".join(sorted(agreed)) if agreed else ", ".join(sorted(manifest_tools)),
                "all_tools": ", ".join(sorted(all_union)),
                "confidence": "high",
                "conflict_detail": "",
            }
        else:
            if ci_union or jenkinsfile_tools:
                ci_str = ", ".join(sorted(ci_union | jenkinsfile_tools))
                manifest_str = ", ".join(sorted(manifest_tools))
                return {
                    "primary_tool": ", ".join(sorted(manifest_tools)),
                    "all_tools": ", ".join(sorted(all_union)),
                    "confidence": "conflict",
                    "conflict_detail": f"Manifest: {manifest_str} | CI: {ci_str}",
                }
            else:
                return {
                    "primary_tool": ", ".join(sorted(manifest_tools)),
                    "all_tools": ", ".join(sorted(all_union)),
                    "confidence": "medium",
                    "conflict_detail": "Sadece manifest var, hiç CI kaynağı yok (build edilmiyor olabilir)",
                }
    
    # JENKINSFILE VAR (servis reposunda)
    if jenkinsfile_tools:
        if tools_intersect(jenkinsfile_tools, ci_union):
            agreed = set()
            for jt in jenkinsfile_tools:
                jn = normalize_tool_name(jt)
                for ct in ci_union:
                    if normalize_tool_name(ct) == jn:
                        agreed.add(jt)
                        break
            return {
                "primary_tool": ", ".join(sorted(agreed)) if agreed else ", ".join(sorted(jenkinsfile_tools)),
                "all_tools": ", ".join(sorted(all_union)),
                "confidence": "high",
                "conflict_detail": "",
            }
        return {
            "primary_tool": ", ".join(sorted(jenkinsfile_tools)),
            "all_tools": ", ".join(sorted(all_union)),
            "confidence": "medium",
            "conflict_detail": "Sadece servis reposundaki Jenkinsfile'dan tespit",
        }
    
    # Sadece CI kaynakları
    triple = role_tools & config_tools & log_tools
    if triple:
        return {
            "primary_tool": ", ".join(sorted(triple)),
            "all_tools": ", ".join(sorted(ci_union)),
            "confidence": "high",
            "conflict_detail": "",
        }
    
    pairs = (role_tools & log_tools) | (role_tools & config_tools) | (config_tools & log_tools)
    if pairs:
        return {
            "primary_tool": ", ".join(sorted(pairs)),
            "all_tools": ", ".join(sorted(ci_union)),
            "confidence": "medium",
            "conflict_detail": (f"Tüm tools: {', '.join(sorted(ci_union))}"
                                if len(ci_union) > len(pairs) else ""),
        }
    
    return {
        "primary_tool": ", ".join(sorted(ci_union)),
        "all_tools": ", ".join(sorted(ci_union)),
        "confidence": "low",
        "conflict_detail": "Sadece tek CI kaynağından tespit",
    }


def correlate(jobs_data, roles_data, repos_data=None, artifactory_data=None):
    rows = []
    repos_data = repos_data or {}
    
    for job in jobs_data:
        invoked_roles = job.get("ansible_roles_invoked", set())
        if isinstance(invoked_roles, list):
            invoked_roles = set(invoked_roles)
        
        # Ansible role tools
        role_tools = set()
        role_languages = set()
        for role_name in invoked_roles:
            role_info = roles_data.get(role_name)
            if role_info and role_info.get("is_build_role"):
                role_tools.update(role_info.get("build_tools", set()))
                role_languages.update(role_info.get("languages", set()))
        
        scm_url = job.get("primary_scm_url", "")
        repo_info = repos_data.get(scm_url, {}) if scm_url else {}
        
        manifest_tools = set(repo_info.get("build_tools_from_manifest", set())) if repo_info else set()
        jenkinsfile_tools = set(repo_info.get("build_tools_from_jenkinsfile", set())) if repo_info else set()
        manifest_languages = set(repo_info.get("languages_from_manifest", set())) if repo_info else set()
        jenkinsfile_languages = set(repo_info.get("languages_from_jenkinsfile", set())) if repo_info else set()
        
        config_tools = set(job.get("build_tools_from_config", set()))
        log_tools = set(job.get("build_tools_from_log", set()))
        
        reconciled = reconcile_with_manifest(
            role_tools=role_tools,
            config_tools=config_tools,
            log_tools=log_tools,
            manifest_tools=manifest_tools,
            jenkinsfile_tools=jenkinsfile_tools,
        )
        
        all_langs = (role_languages | set(job.get("languages_from_config", set()))
                     | set(job.get("languages_from_log", set()))
                     | manifest_languages | jenkinsfile_languages)
        
        # Primary language: manifest > jenkinsfile > diğer (Wrapper hariç)
        non_wrapper = all_langs - {"Wrapper", "Generic publish"}
        if manifest_languages:
            primary_lang = sorted(manifest_languages)[0]
        elif jenkinsfile_languages - {"Wrapper", "Generic publish"}:
            primary_lang = sorted(jenkinsfile_languages - {"Wrapper", "Generic publish"})[0]
        elif non_wrapper:
            primary_lang = sorted(non_wrapper)[0]
        else:
            primary_lang = sorted(all_langs)[0] if all_langs else ""
        
        manifests_found = repo_info.get("manifests_found", []) if repo_info else []
        
        rows.append({
            "service_repo_url": scm_url,
            "platform": detect_platform(scm_url),
            "primary_language": primary_lang,
            "primary_build_tool": reconciled["primary_tool"],
            "confidence": reconciled["confidence"],
            "all_languages": ", ".join(sorted(non_wrapper)),
            "all_build_tools": reconciled["all_tools"],
            "manifests_found": ", ".join(manifests_found[:6]),
            "has_service_jenkinsfile": repo_info.get("has_jenkinsfile", False) if repo_info else False,
            "tools_from_manifest": ", ".join(sorted(manifest_tools)),
            "tools_from_service_jenkinsfile": ", ".join(sorted(jenkinsfile_tools)),
            "tools_from_role": ", ".join(sorted(role_tools)),
            "tools_from_jenkins_config": ", ".join(sorted(config_tools)),
            "tools_from_jenkins_log": ", ".join(sorted(log_tools)),
            "ansible_roles_invoked": ", ".join(sorted(invoked_roles)),
            "uses_ansible": (job.get("uses_ansible", False) or
                             (repo_info.get("jenkinsfile_uses_ansible", False) if repo_info else False)),
            "jenkins_job": job["job_full_name"],
            "jenkins_url": job["job_url"],
            "definition_type": job.get("definition_type", job.get("job_class", "")),
            "branches": ", ".join(job.get("branches", [])[:3]),
            "config_fetched": job.get("config_fetched", False),
            "log_fetched": job.get("log_fetched", False),
            "repo_scanned": bool(repo_info),
            "conflict_detail": reconciled["conflict_detail"],
        })
    
    return rows


def write_excel(output_path, correlated_rows, roles_data, jobs_data, repos_data=None):
    repos_data = repos_data or {}
    wb = Workbook()
    wb.remove(wb.active)
    
    df_main = pd.DataFrame(correlated_rows)
    if not df_main.empty:
        priority = [
            "service_repo_url", "platform", "primary_language", "primary_build_tool",
            "confidence", "manifests_found", "has_service_jenkinsfile",
            "ansible_roles_invoked", "uses_ansible",
            "all_languages", "all_build_tools",
            "tools_from_manifest", "tools_from_service_jenkinsfile",
            "tools_from_role", "tools_from_jenkins_config", "tools_from_jenkins_log",
            "jenkins_job", "definition_type", "branches",
            "jenkins_url", "config_fetched", "log_fetched", "repo_scanned",
            "conflict_detail",
        ]
        df_main = df_main[[c for c in priority if c in df_main.columns]]
    _write_sheet(wb, "1. Cross-Reference (Ana)", df_main)
    
    if not df_main.empty:
        conf_summary = df_main.groupby("confidence").size().reset_index(name="count").sort_values("count", ascending=False)
    else:
        conf_summary = pd.DataFrame()
    _write_sheet(wb, "2. Confidence Dağılımı", conf_summary)
    
    if not df_main.empty:
        conflicts = df_main[df_main["confidence"].isin(["conflict", "low", "no-data"])]
        cols_show = ["service_repo_url", "confidence", "primary_build_tool",
                     "tools_from_manifest", "tools_from_service_jenkinsfile",
                     "tools_from_role", "tools_from_jenkins_config", "tools_from_jenkins_log",
                     "ansible_roles_invoked", "jenkins_job", "conflict_detail"]
        conflicts_view = conflicts[[c for c in cols_show if c in conflicts.columns]] if not conflicts.empty else pd.DataFrame()
    else:
        conflicts_view = pd.DataFrame()
    _write_sheet(wb, "3. İncelenecek (Conflict)", conflicts_view)
    
    repo_rows = []
    for url, info in sorted(repos_data.items()):
        if not isinstance(info, dict): continue
        meta_parts = []
        for fname, meta in info.get("manifest_metadata", {}).items():
            if fname == "pom.xml":
                gid = meta.get("groupId", "")
                aid = meta.get("artifactId", "")
                if gid or aid:
                    meta_parts.append(f"maven:{gid}:{aid}")
            elif fname == "package.json":
                n = meta.get("name", "")
                if n: meta_parts.append(f"npm:{n}")
            elif fname == "Chart.yaml":
                n = meta.get("name", "")
                if n: meta_parts.append(f"helm:{n}")
        repo_rows.append({
            "service_repo_url": url,
            "platform": info.get("platform", ""),
            "manifests_found": ", ".join(info.get("manifests_found", [])),
            "has_jenkinsfile": info.get("has_jenkinsfile", False),
            "jenkinsfile_uses_ansible": info.get("jenkinsfile_uses_ansible", False),
            "build_tools_from_manifest": ", ".join(sorted(info.get("build_tools_from_manifest", []))),
            "build_tools_from_jenkinsfile": ", ".join(sorted(info.get("build_tools_from_jenkinsfile", []))),
            "metadata": " | ".join(meta_parts),
            "errors": ", ".join(info.get("errors", [])),
        })
    df_repos = pd.DataFrame(repo_rows)
    _write_sheet(wb, "4. Servis Repo Analizi", df_repos)
    
    role_rows = [{
        "role_name": name,
        "is_build_role": info.get("is_build_role", False),
        "build_tools": ", ".join(sorted(info.get("build_tools", set()))),
        "languages": ", ".join(sorted(info.get("languages", set()))),
        "sample_command": (info.get("snippets", [""])[0] if info.get("snippets") else ""),
    } for name, info in sorted(roles_data.items())]
    _write_sheet(wb, "5. Ansible Roles", pd.DataFrame(role_rows))
    
    tech_counter = Counter()
    for row in correlated_rows:
        lang = row.get("primary_language", "")
        if lang and lang not in ("Wrapper", "Generic publish"):
            tech_counter[lang] += 1
    df_tech = pd.DataFrame([{"language": k, "service_count": v} for k, v in tech_counter.most_common()])
    _write_sheet(wb, "6. Teknoloji Dağılımı", df_tech)
    
    LANG_TO_PKG = {"Java/Kotlin": "maven", ".NET": "nuget", "Node.js": "npm",
                   "Python": "pypi", "Container": "docker", "K8s Chart": "helm",
                   "Go": "go", "Ruby": "gems", "PHP": "composer", "Rust": "cargo"}
    detected_langs = {row.get("primary_language") for row in correlated_rows if row.get("primary_language")}
    suggestions = []
    seen = set()
    for lang in sorted(detected_langs):
        pkg = LANG_TO_PKG.get(lang)
        if not pkg or pkg in seen: continue
        seen.add(pkg)
        suggestions.append({
            "package_type": pkg, "language": lang,
            "release_repo": f"{pkg}-release-local",
            "snapshot_repo": f"{pkg}-snapshot-local" if pkg == "maven" else "(N/A)",
            "remote_repo": f"{pkg}-remote", "virtual_repo": f"{pkg}-virtual",
        })
    _write_sheet(wb, "7. Yeni Repo Önerileri", pd.DataFrame(suggestions))
    
    jobs_simple = [{
        "job_full_name": j["job_full_name"], "job_class": j["job_class"],
        "primary_scm_url": j.get("primary_scm_url", ""),
        "uses_ansible": j.get("uses_ansible", False),
        "ansible_roles": ", ".join(sorted(j.get("ansible_roles_invoked", set()) if isinstance(j.get("ansible_roles_invoked"), set) else j.get("ansible_roles_invoked", []))),
        "config_fetched": j.get("config_fetched", False),
        "log_fetched": j.get("log_fetched", False),
    } for j in jobs_data]
    _write_sheet(wb, "8. Ham Jenkins Jobs", pd.DataFrame(jobs_simple))
    
    wb.save(output_path)
    print(f"  ✅ Excel: {output_path}")


def _write_sheet(wb, name, df):
    ws = wb.create_sheet(name)
    if df.empty:
        ws["A1"] = "(boş)"
        return
    
    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    header_fill = PatternFill("solid", start_color="1F4E78")
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    
    for ci, cn in enumerate(df.columns, 1):
        c = ws.cell(row=1, column=ci, value=cn)
        c.font = header_font; c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    
    confidence_idx = (list(df.columns).index("confidence") + 1) if "confidence" in df.columns else None
    
    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            display = ", ".join(sorted(val)) if isinstance(val, set) else val
            c = ws.cell(row=ri, column=ci, value=display)
            c.font = Font(name="Arial", size=10); c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if ci == confidence_idx:
                if val == "high":
                    c.fill = PatternFill("solid", start_color="C6EFCE")
                elif val == "medium":
                    c.fill = PatternFill("solid", start_color="FFEB9C")
                elif val == "conflict":
                    c.fill = PatternFill("solid", start_color="F4B084")
                elif val in ("low", "no-data"):
                    c.fill = PatternFill("solid", start_color="FFC7CE")
    
    for ci, cn in enumerate(df.columns, 1):
        col_values = df.iloc[:, ci-1].astype(str).tolist()[:200]
        max_len = max([len(str(cn))] + [len(v) for v in col_values])
        ws.column_dimensions[get_column_letter(ci)].width = min(max(max_len, 12) + 2, 50)
    
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
